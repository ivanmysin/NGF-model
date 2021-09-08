"""
Версия, содержащая 1 пирамидный нейрон, 1 генератор импульсов, 
множество NGF нейронов.
Организация: 
    По 1-2 NGF на поток с rank > 1 
    В потоке с rank = 1 пирамидный нейрон + генератор импульсов
    В потоке с rank = 0 сбор данных из других потоков
ВНИМАНИЕ! Щелевой контакт должен быть отдельным для каждого соединения
(в отличие от синапсов).
Вопрос: должен ли быть уникальный компартмент? Или можно бесконечное кол-во
щелевых контактов напихать в одну секцию? - Вроде да.
Для работы t.record(n.h._ref_t) в потоке должны быть секции
Доделать НМДА синапсы
НЕ ЗАБЫТЬ выставить температуру! (n.h.celsius = 37)
- Температура не решает проблем с подпороговой динамикой 
пирамидных нейронов, возможно дело в nseg (нет), дело в 
h-токах (в нейроне Turi), в нейроне Кутсуридиса отклбчение 
всех проводимостей не решает проблемы, кривая утечка? кальций?
- У нейрона Turi странная надпороговая динамика, как будто нет
рефрактерного периода. Также синапсы, быстро достигают насыщения 
по весу, вызывают пачки ПД только вместе (AMPA + NMDA),при
этом веса надо снижать (!). В общем бред!
- Несколько n.load_mechanisms могут конкурировать друг с 
другом и вызывать ошибку.
- В пирамиде Кутсуридиса если полностью убрать все токи, кроме 
натриевых, то после ПД все равно генерируется следовуха
"""
from mpi4py import MPI
comm = MPI.COMM_WORLD
rank = comm.Get_rank()
#print('size ', comm.Get_size())
print('RANK----------------------', rank)

import neuron as n
from neuron.units import ms, mV
n.h.load_file("stdrun.hoc")
n.h.load_file("stdgui.hoc")
n.h.load_file("import3d.hoc")

import numpy as np
import math
import random
import matplotlib.pyplot as ppl
#import presimulation_lib as pre_lib

n.load_mechanisms(r"C:\Users\Povarov\Mechs\NGF_Bezaire")
#n.load_mechanisms(r"C:\Users\Povarov\Mechs\Turi_mech")
n.load_mechanisms(r"C:\Users\Povarov\Mechs\Hoc_Mod")
#n.load_mechanisms(r"C:\Users\Povarov\Mechs\MyExp2Syn")
n.load_mechanisms(r"C:\Users\Povarov\Mechs\gap")
n.load_mechanisms(r"C:\Users\Povarov\Mechs\ArtCells")

import BezaireNgf
n.h.load_file("CA1PC.hoc")
#n.h.load_file("Turi_PyrCell.hoc")
 
NgfPerRank = 1 
ngf_num = 2
#########################
    #INITIALIZATION
#########################
pc = n.h.ParallelContext()
CellLists = {'NgfList':n.h.List(), 'PyrList':n.h.List(), 'ArtList': n.h.List()}
GapList = n.h.List()

ListKey = ''

if rank == 1:
    ListKey = 'PyrList'
    Pyr = n.h.CA1PyramidalCell(rank, 0)#Cutsuridis PC
    #Pyr = n.h.PyramidalCell(rank, 0) #Turi PC
    CellLists.get(ListKey).append(Pyr)
    pc.set_gid2node(rank, pc.id())
    ProtoCon = n.h.NetCon(CellLists.get(ListKey)[0].soma(0.5)._ref_v, 
               None, sec = CellLists.get(ListKey)[0].soma)
    ProtoCon.threshold = -30 #* mV
    pc.cell(rank, ProtoCon)
    import presimulation_lib as pre_lib
    ListKey = 'ArtList'
    AtrCell = n.h.ArtificialRhytmicCell(0, 0)
    theta_kappa, theta_i0 = pre_lib.r2kappa(0.9)
    AtrCell.kappa = theta_kappa
    AtrCell.I0 = theta_i0
    AtrCell.spike_rate = 5
    CellLists.get(ListKey).append(AtrCell)
    #КАК ПОДКЛЮЧИТЬ ГЕНЕРАТОР ЧЕРЕЗ ГИД?
    art_gid = 0
    pc.set_gid2node(art_gid, pc.id())
    for art_con in range(1):
        ProtoCon_art = n.h.NetCon(CellLists.get(ListKey)[0], None)
        pc.cell(art_gid, ProtoCon_art)
    print('Art + Pyr init  done')

if rank > 1:
    for neur in range(NgfPerRank):
        gid = neur + 1 + ((rank - 1) * NgfPerRank)
        print('rank ', rank, ' gid ', gid)
        ListKey = 'NgfList'
        Ngf = BezaireNgf.BezaireNgf(gid) 
        CellLists.get(ListKey).append(Ngf)
        pc.set_gid2node(gid, pc.id())
        ProtoCon = n.h.NetCon(CellLists.get(ListKey)[neur].AllSections[0](0.5)._ref_v, 
                   None, sec = CellLists.get(ListKey)[neur].AllSections[0])
        ProtoCon.threshold = -30 #* mV
        pc.cell(gid, ProtoCon)
        for i in range(ngf_num):
            gap_num = gid * 10 + i
            print('for gid ', gid, ' gap_num ', gap_num)
            pc.source_var(CellLists.get(ListKey)[neur].AllSections[0](0.5)._ref_v, 
                         gap_num, sec = CellLists.get(ListKey)[neur].AllSections[0])
            gap = n.h.gap3(CellLists.get(ListKey)[neur].AllSections[2](0.5))
            GapList.append(gap)
        if rank == 2:
            iclamp = n.h.IClamp(CellLists.get(ListKey)[0].AllSections[0](0.5))
            iclamp.amp = 0.2 # nA
            iclamp.delay = 120 # ms
            iclamp.dur = 20 # ms 
    print('Ngf + Gap    done')
    print('GapList ', len(GapList))

############################
    #CONNECTION
############################
NgfVList = n.h.List()
for i in range(NgfPerRank):
    Ngf_v = n.h.Vector()
    NgfVList.append(Ngf_v)
PyrVList = n.h.List()
Pyr_v = n.h.Vector()
PyrVList.append(Pyr_v)
t = n.h.Vector()

NgfToNgf = n.h.List()
NgfToPyr = n.h.List()
ArtToPyr_AMPA = n.h.List()
ArtToPyr_NMDA = n.h.List()
ArtToNgf = n.h.List()
print('Output lists init   done')

if rank == 1:
    ListKey = 'PyrList'
    """
    for pre_gid in range(2, ngf_num + 3, 1): #Входы от NGF на пирамидный нейрон
        conn = pc.gid_connect(pre_gid, CellLists.get(ListKey)[0].pre_list[23])
        conn.delay = 20 * (rank + 1)
        conn.weight[0] = 1
        conn.threshold = -30 * mV 
        NgfToPyr.append(conn)
    """
    PyrVList[0].record(CellLists.get(ListKey)[0].soma(0.5)._ref_v)
    #ListKey = 'ArtList'
    for art_con in range(1): #Несколько входов на одну клетку
        conn = pc.gid_connect(0, CellLists.get(ListKey)[0].pre_list[23]) #23 for cutsuridis #1
        conn.weight[0] = 0.0 #Двухфазный ВПСП, это из-за h-токов
        ArtToPyr_AMPA.append(conn)
        conn = pc.gid_connect(0, CellLists.get(ListKey)[0].pre_list[55]) #55 for cutsuridis #7
        conn.weight[0] = 0.0 #Синапсы меняются вместе с типом нейрона!
        ArtToPyr_NMDA.append(conn)
    
    iclamp_pyr = n.h.IClamp(CellLists.get(ListKey)[0].soma(0.5))
    iclamp_pyr.amp = 0.4 # nA
    iclamp_pyr.delay = 120 # ms
    iclamp_pyr.dur = 200 # ms
    """
    syn = n.h.MyExp2Syn(0.5, sec = CellLists.get(ListKey)[0].soma)
    syn.tau1 = 0.5
    syn.tau2 = 3
    syn.e = 0
    conn = pc.gid_connect(0, syn) #23 for cutsuridis #1
    conn.weight[0] = 0.04 #Двухфазный ВПСП, это из-за h-токов
    ArtToPyr_AMPA.append(conn)
    """
    t.record(n.h._ref_t)
    print('Art + Pyr conns........................done')
  
if rank > 1:
    for neur in range(NgfPerRank):
        ListKey = 'NgfList'
        gid = CellLists.get(ListKey)[neur].gid
        for pre_gid in range(2, ngf_num + 2, 1):
            if pre_gid != gid: #Без автоиннервации
                conn = pc.gid_connect(pre_gid, CellLists.get(ListKey)[neur].SynList_ex[0])
                conn.delay = 20 * (rank + 1)
                conn.weight[0] = 1
                conn.threshold = -30 * mV 
                NgfToNgf.append(conn)
                #Может возникнуть конфликт - множество v сходятся на одном vgap:
                gap_num = pre_gid * 10 + pre_gid - 2
                print('for ngf_num ', neur, ' gap_num ', gap_num)
                #print('pre_gid - 2 ', pre_gid - 2)
                pc.target_var(GapList[pre_gid - 2], GapList[pre_gid - 2]._ref_vgap, gap_num)
                GapList[pre_gid - 2].g = 0.17
        #pc.setup_transfer()
        NgfVList[neur].record(CellLists.get(ListKey)[neur].AllSections[0](0.5)._ref_v)
        #Проекция генератора импульсов на все NGF
        conn = pc.gid_connect(0, CellLists.get(ListKey)[neur].SynList_ex[0])
        #Почему такой маленький вес нужен по сравнению с пирамидой?
        conn.weight[0] = 0 #0.001
        ArtToNgf.append(conn)
        t.record(n.h._ref_t)
        print('Ngf + Gap conns    done')
####################################
    #SIMULATION AND DATA COLLECTION
####################################
pc.setup_transfer()

n.h.tstop = 1000 * ms
pc.set_maxstep(5 * ms)
n.h.celsius = 37
n.h.finitialize()
pc.psolve(1000 * ms) #замена функции run()
pc.barrier()

pc.done() #тормозим все хосты
print('DONE')
if rank > 1:
    ppl.plot(t, NgfVList[0])
    ppl.show()
if rank == 1:
    ppl.plot(t, PyrVList[0])
    ppl.title('Pyr')
    ppl.show()
"""
NgfVList_send = []
for vec in NgfVList:
    NgfVList_send.append(list(vec))
PyrVList_send = []
for vec in PyrVList:
    PyrVList_send.append(vec) 
#t = list(t)
NgfSum = comm.gather(NgfVList_send, root = 0)
PyrSum = comm.gather(PyrVList_send, root = 0)
if rank == 0:
    import openpyxl as xl
    book = xl.Workbook()
    sheet = book.active
    #sheet.append(t)
    for i in PyrSum:
        sheet.append(i)
    for i in NgfSum:
        sheet.append(i)
    book.save('aaa.xlsx')
    book.close()
    #f = open(r"D:\Python_ver_3\Pyt_last\WPy64-3740\settings\.spyder-py3\SimResult\TestRes.txt", 'w')
    #f.write(str(NgfSum))
    #f.close()
"""